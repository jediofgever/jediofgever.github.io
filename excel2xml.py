import openpyxl

from xml.dom  import minidom
import os 

wb =openpyxl.load_workbook('training.xlsx')
sheet = wb.active

count =1


while count<8145:
	root = minidom.Document()


	xml = root.createElement('annotation')
	root.appendChild(xml)

	productChild = root.createElement('folder')
	productChild.appendChild(root.createTextNode('images'))
	xml.appendChild(productChild)

	productChild = root.createElement('filename')
	productChild.appendChild(root.createTextNode(str(count).zfill(5)+'.jpg'))
	xml.appendChild(productChild)

	productChild = root.createElement('path')
	productChild.appendChild(root.createTextNode('/home/atas/Desktop/images/'+str(count).zfill(5)+'.jpg'))
	xml.appendChild(productChild)



	productChild = root.createElement('source')
	childOfProduct = root.createElement('database')
	childOfProduct.appendChild(root.createTextNode('Unknown'))
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)


	productChild = root.createElement('size')
	childOfProduct = root.createElement('width')
	childOfProduct.appendChild(root.createTextNode('640'))
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)

	childOfProduct = root.createElement('height')
	childOfProduct.appendChild(root.createTextNode('480'))
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)

	childOfProduct = root.createElement('depth')
	childOfProduct.appendChild(root.createTextNode('3'))
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)


	productChild = root.createElement('segmented')
	productChild.appendChild(root.createTextNode('0'))
	xml.appendChild(productChild)

	productChild = root.createElement('object')
	childOfProduct = root.createElement('name')
	childOfProduct.appendChild(root.createTextNode('vehicle'))
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)

	childOfProduct = root.createElement('pose')
	childOfProduct.appendChild(root.createTextNode('Unspecified'))
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)

	childOfProduct = root.createElement('truncated')
	childOfProduct.appendChild(root.createTextNode('0'))
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)

	childOfProduct = root.createElement('difficult')
	childOfProduct.appendChild(root.createTextNode('0'))
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)


	childOfProduct = root.createElement('bndbox')

        subChildOfProduct = root.createElement('xmin')
	subChildOfProduct.appendChild(root.createTextNode(str(sheet.cell(row=count,column=5).value)))
	childOfProduct.appendChild(subChildOfProduct)
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)


	subChildOfProduct = root.createElement('ymin')
	subChildOfProduct.appendChild(root.createTextNode(str(sheet.cell(row=count,column=4).value)))
	childOfProduct.appendChild(subChildOfProduct)
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)

	subChildOfProduct = root.createElement('xmax')
	subChildOfProduct.appendChild(root.createTextNode(str(sheet.cell(row=count,column=3).value)))
	childOfProduct.appendChild(subChildOfProduct)
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)

	subChildOfProduct = root.createElement('ymax')
	subChildOfProduct.appendChild(root.createTextNode(str(sheet.cell(row=count,column=2).value)))
	childOfProduct.appendChild(subChildOfProduct)
	productChild.appendChild(childOfProduct)
	xml.appendChild(productChild)

	     
 



	xml_str = root.toprettyxml(indent="\t")



	save_path_file = str(count).zfill(5)+'.xml'
        with open(save_path_file,"w") as f:
	       
             f.write(xml_str)
            


        count += 1



for x in range (1,8000):
    print(sheet.cell(row=x,column=3).value)





